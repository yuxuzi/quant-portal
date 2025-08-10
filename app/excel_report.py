"""
Excel Report Library - Structured library for creating Excel reports with native formatting,
DataFrames, and embedded Plotly charts.
Optimized for quantitative finance teams using openpyxl and plotly/kaleido stack.

Version: 1.0.0
"""

import os
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Self, Tuple

import numpy as np
import pandas as pd
from pydantic import BaseModel, Field, model_validator

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import plotly.express as px
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False


class ExcelStyle(BaseModel):
    """Excel cell styling configuration for quantitative reporting."""
    font_name: str = "Calibri"
    font_size: int = 11
    font_bold: bool = False
    font_color: str = "000000"  # Black
    fill_color: str | None = None  # Hex color without #
    border_style: str = "thin"  # thin, medium, thick
    border_color: str = "000000"
    alignment_horizontal: str = "general"  # left, center, right, general
    alignment_vertical: str = "bottom"  # top, center, bottom
    number_format: str = "General"

    model_config = {"arbitrary_types_allowed": True, "extra": "forbid"}


class TableStyle(BaseModel):
    """Predefined table styling configurations optimized for quant finance reports."""
    header: ExcelStyle = Field(default_factory=lambda: ExcelStyle(
        font_bold=True, fill_color="D9E1F2", alignment_horizontal="center"
    ))
    data: ExcelStyle = Field(default_factory=lambda: ExcelStyle())
    alternating_row: ExcelStyle = Field(default_factory=lambda: ExcelStyle(fill_color="F2F2F2"))

    model_config = {"arbitrary_types_allowed": True, "extra": "forbid"}

    @classmethod
    def financial_table(cls) -> Self:
        """Financial/trading table style with professional blue theme."""
        return cls(
            header=ExcelStyle(font_bold=True, fill_color="4472C4", font_color="FFFFFF", 
                              alignment_horizontal="center"),
            data=ExcelStyle(alignment_horizontal="right", number_format="#,##0.00"),
            alternating_row=ExcelStyle(fill_color="F8F9FA")
        )
    
    @classmethod
    def risk_table(cls) -> Self:
        """Risk metrics table style with alert red theme."""
        return cls(
            header=ExcelStyle(font_bold=True, fill_color="E74C3C", font_color="FFFFFF",
                              alignment_horizontal="center"),
            data=ExcelStyle(alignment_horizontal="center", number_format="0.00%"),
            alternating_row=ExcelStyle(fill_color="FADBD8")
        )
    
    @classmethod
    def pnl_table(cls) -> Self:
        """P&L table style with green profit theme."""
        return cls(
            header=ExcelStyle(font_bold=True, fill_color="28A745", font_color="FFFFFF",
                              alignment_horizontal="center"),
            data=ExcelStyle(alignment_horizontal="right", number_format='"$"#,##0.00;[Red]("$"#,##0.00)'),
            alternating_row=ExcelStyle(fill_color="E8F5E8")
        )

    @model_validator(mode='after')
    def validate_styles(self) -> Self:
        """Validate style configurations."""
        if self.header.font_size < 8 or self.header.font_size > 72:
            raise ValueError("Header font size must be between 8 and 72")
        return self


class ExcelComponent:
    """Base class for Excel report components."""
    
    def __init__(self, title: str | None = None, start_row: int = 1, start_col: int = 1) -> None:
        """
        Initialize Excel component.
        
        Args:
            title: Component title
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
        """
        self.title = title
        self.start_row = start_row
        self.start_col = start_col
        self.end_row = start_row
        self.end_col = start_col
        
    def write_to_worksheet(self, worksheet, current_row: int | None = None) -> int:
        """Write component to worksheet and return next available row."""
        raise NotImplementedError("Subclasses must implement write_to_worksheet")
        
    def _apply_style(self, cell, style: ExcelStyle) -> None:
        """Apply ExcelStyle to a cell."""
        if not HAS_OPENPYXL:
            return
            
        # Font styling
        cell.font = Font(
            name=style.font_name,
            size=style.font_size,
            bold=style.font_bold,
            color=style.font_color
        )
        
        # Fill styling
        if style.fill_color:
            cell.fill = PatternFill(start_color=style.fill_color, end_color=style.fill_color, fill_type="solid")
            
        # Border styling
        border_side = Side(style=style.border_style, color=style.border_color)
        cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Alignment
        cell.alignment = Alignment(
            horizontal=style.alignment_horizontal,
            vertical=style.alignment_vertical
        )
        
        # Number format
        cell.number_format = style.number_format


class TextComponent(ExcelComponent):
    """Component for text content in Excel."""
    
    def __init__(self, text: str, style: ExcelStyle | None = None, **kwargs) -> None:
        """
        Initialize text component.
        
        Args:
            text: Text content
            style: Excel styling
            **kwargs: Additional component arguments
        """
        super().__init__(**kwargs)
        self.text = text
        self.style = style or ExcelStyle(font_size=12)
        
    def write_to_worksheet(self, worksheet, current_row: int | None = None) -> int:
        """Write text to worksheet."""
        row = current_row or self.start_row
        
        # Add title if provided
        if self.title:
            title_cell = worksheet.cell(row=row, column=self.start_col, value=self.title)
            title_style = ExcelStyle(font_bold=True, font_size=14)
            self._apply_style(title_cell, title_style)
            row += 2
            
        # Add text content
        text_cell = worksheet.cell(row=row, column=self.start_col, value=self.text)
        self._apply_style(text_cell, self.style)
        
        self.end_row = row
        return row + 2  # Return next available row with spacing


class DataFrameComponent(ExcelComponent):
    """Component for DataFrames with native Excel formatting, optimized for quant data."""
    
    def __init__(self, df: pd.DataFrame, 
                 table_style: TableStyle | None = None,
                 number_formats: dict[str, str] | None = None,
                 conditional_formatting: dict[str, dict[str, Any]] | None = None,
                 freeze_header: bool = True,
                 autofit_columns: bool = True,
                 **kwargs) -> None:
        """
        Initialize DataFrame component.
        
        Args:
            df: pandas DataFrame
            table_style: TableStyle configuration
            number_formats: Column-specific number formatting {'col': 'format'}
            conditional_formatting: Conditional formatting rules
            freeze_header: Whether to freeze header row
            autofit_columns: Whether to auto-fit column widths
            **kwargs: Additional component arguments
        """
        super().__init__(**kwargs)
        self.df = df
        self.table_style = table_style or TableStyle()
        self.number_formats = number_formats or {}
        self.conditional_formatting = conditional_formatting or {}
        self.freeze_header = freeze_header
        self.autofit_columns = autofit_columns
        
    def write_to_worksheet(self, worksheet, current_row: int | None = None) -> int:
        """Write DataFrame to worksheet with native Excel formatting."""
        start_row = current_row or self.start_row
        
        # Add title if provided
        if self.title:
            title_cell = worksheet.cell(row=start_row, column=self.start_col, value=self.title)
            title_style = ExcelStyle(font_bold=True, font_size=14)
            self._apply_style(title_cell, title_style)
            start_row += 2
            
        # Write DataFrame to worksheet
        df_start_row = start_row
        
        # Write headers
        for col_idx, column in enumerate(self.df.columns, start=0):
            cell = worksheet.cell(row=df_start_row, column=self.start_col + col_idx, value=column)
            self._apply_style(cell, self.table_style.header)
            
        # Write data rows
        for row_idx, row_data in enumerate(self.df.itertuples(index=False), start=0):
            excel_row = df_start_row + 1 + row_idx
            
            # Determine if alternating row
            use_alt_style = row_idx % 2 == 1 and self.table_style.alternating_row.fill_color is not None
            row_style = self.table_style.alternating_row if use_alt_style else self.table_style.data
            
            for col_idx, value in enumerate(row_data, start=0):
                cell = worksheet.cell(row=excel_row, column=self.start_col + col_idx, value=value)
                
                # Apply base style
                self._apply_style(cell, row_style)
                
                # Apply column-specific number formatting
                col_name = self.df.columns[col_idx]
                if col_name in self.number_formats:
                    cell.number_format = self.number_formats[col_name]
                    
        # Apply conditional formatting
        self._apply_conditional_formatting(worksheet, df_start_row)
        
        # Auto-fit columns
        if self.autofit_columns:
            self._autofit_columns(worksheet)
            
        # Freeze header
        if self.freeze_header:
            worksheet.freeze_panes = worksheet.cell(row=df_start_row + 1, column=1)
            
        self.end_row = df_start_row + len(self.df)
        return self.end_row + 3  # Return next available row with spacing
        
    def _apply_conditional_formatting(self, worksheet, header_row: int) -> None:
        """Apply conditional formatting rules for quant metrics."""
        if not self.conditional_formatting:
            return
            
        data_start_row = header_row + 1
        data_end_row = header_row + len(self.df)
        
        for col_name, formatting_rule in self.conditional_formatting.items():
            if col_name not in self.df.columns:
                continue
                
            col_idx = self.df.columns.get_loc(col_name)
            col_letter = get_column_letter(self.start_col + col_idx)
            cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
            
            rule_type = formatting_rule.get('type')
            
            if rule_type == 'color_scale':
                # Color scale formatting (red-yellow-green) for risk levels
                rule = ColorScaleRule(
                    start_type='min', start_color='FF6B6B',  # Red for high risk
                    mid_type='percentile', mid_value=50, mid_color='FFE66D',  # Yellow
                    end_type='max', end_color='4ECDC4'  # Green for low risk
                )
                worksheet.conditional_formatting.add(cell_range, rule)
                
            elif rule_type == 'data_bars':
                # Data bars for performance metrics
                rule = DataBarRule(
                    start_type='min', end_type='max',
                    color='4472C4', showValue=True
                )
                worksheet.conditional_formatting.add(cell_range, rule)
                
            elif rule_type == 'positive_negative':
                # Color positive values green, negative red for P&L
                pos_rule = CellIsRule(operator='greaterThan', formula=['0'], 
                                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                                      font=Font(color='006100'))
                neg_rule = CellIsRule(operator='lessThan', formula=['0'],
                                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                                      font=Font(color='9C0006'))
                worksheet.conditional_formatting.add(cell_range, pos_rule)
                worksheet.conditional_formatting.add(cell_range, neg_rule)
                
    def _autofit_columns(self, worksheet) -> None:
        """Auto-fit column widths based on content for better readability."""
        for col_idx, column in enumerate(self.df.columns, start=0):
            column_letter = get_column_letter(self.start_col + col_idx)
            
            # Calculate max width needed
            max_length = len(str(column))  # Header length
            for value in self.df.iloc[:, col_idx]:
                max_length = max(max_length, len(str(value)))
                
            # Set column width (with some padding, max 50 for quant data)
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)


class PlotlyChartComponent(ExcelComponent):
    """Component for embedding Plotly charts as images in Excel, suitable for quant visualizations."""
    
    def __init__(self, figure: go.Figure | None = None,
                 plot_function: Callable[..., go.Figure] | None = None,
                 plot_args: tuple[Any, ...] = (),
                 plot_kwargs: dict[str, Any] | None = None,
                 image_format: str = "png",
                 width: int = 800,
                 height: int = 500,
                 scale: float = 2.0,
                 **kwargs) -> None:
        """
        Initialize Plotly chart component.
        
        Args:
            figure: Plotly Figure object
            plot_function: Function that returns a Plotly figure
            plot_args: Arguments for plot_function
            plot_kwargs: Keyword arguments for plot_function
            image_format: Image format ('png', 'jpg')
            width: Image width in pixels
            height: Image height in pixels
            scale: Scale factor for high-DPI displays
            **kwargs: Additional component arguments
        """
        super().__init__(**kwargs)
        self.figure = figure
        self.plot_function = plot_function
        self.plot_args = plot_args
        self.plot_kwargs = plot_kwargs or {}
        self.image_format = image_format
        self.width = width
        self.height = height
        self.scale = scale
        
        if not HAS_PLOTLY:
            raise ImportError("Plotly is required for PlotlyChartComponent")
            
    def write_to_worksheet(self, worksheet, current_row: int | None = None) -> int:
        """Write Plotly chart as image to worksheet."""
        start_row = current_row or self.start_row
        
        # Add title if provided
        if self.title:
            title_cell = worksheet.cell(row=start_row, column=self.start_col, value=self.title)
            title_style = ExcelStyle(font_bold=True, font_size=14)
            self._apply_style(title_cell, title_style)
            start_row += 2
            
        # Generate figure if needed
        if self.figure is None:
            if self.plot_function:
                self.figure = self.plot_function(*self.plot_args, **self.plot_kwargs)
            else:
                # Create placeholder text
                placeholder_cell = worksheet.cell(row=start_row, column=self.start_col, 
                                                  value="No chart data available")
                return start_row + 2
                
        # Convert Plotly figure to image bytes
        img_bytes = self.figure.to_image(
            format=self.image_format,
            width=self.width,
            height=self.height,
            scale=self.scale
        )
        
        # Use tempfile for temporary image file
        with tempfile.NamedTemporaryFile(suffix=f".{self.image_format}", delete=False) as temp_file:
            temp_filename = temp_file.name
            temp_file.write(img_bytes)
        
        try:
            # Insert image into worksheet
            img = Image(temp_filename)
            
            # Position image
            cell_address = worksheet.cell(row=start_row, column=self.start_col).coordinate
            img.anchor = cell_address
            
            # Scale image to fit nicely in Excel for quant reports
            img.width = self.width * 0.75
            img.height = self.height * 0.75
            
            worksheet.add_image(img)
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_filename):
                os.remove(temp_filename)
                
        # Calculate rows occupied by image (rough estimate, adjusted for quant charts)
        rows_occupied = max(int(self.height * 0.75 / 18), 20)  # Adjusted for denser charts
        self.end_row = start_row + rows_occupied
        
        return self.end_row + 3  # Increased spacing for readability


class ExcelReport:
    """Main Excel report class that combines components across multiple worksheets for quant analysis."""
    
    def __init__(self, filename: str, title: str | None = None) -> None:
        """
        Initialize Excel report.
        
        Args:
            filename: Output Excel filename
            title: Report title
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for ExcelReport")
            
        self.filename = filename
        self.title = title or "Quantitative Report"
        self.workbook = Workbook()
        self.worksheets: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
        self.components: dict[str, list[ExcelComponent]] = {}
        
        # Remove default worksheet and create summary
        self.workbook.remove(self.workbook.active)
        self._create_summary_worksheet()
        
    def _create_summary_worksheet(self) -> None:
        """Create summary/cover worksheet with quant-specific metadata."""
        summary_ws = self.workbook.create_sheet("Summary", 0)
        self.worksheets["Summary"] = summary_ws
        self.components["Summary"] = []
        
        # Add report title and metadata
        title_component = TextComponent(
            text=self.title,
            style=ExcelStyle(font_size=20, font_bold=True, alignment_horizontal="center"),
            start_row=2
        )
        
        metadata_text = f"""Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        
This report contains quantitative analysis including:
• Performance metrics and P&L breakdown
• Risk analysis and exposure metrics  
• Market data and statistical summaries
• Interactive charts and visualizations
        
Optimized for quantitative finance teams."""
        
        metadata_component = TextComponent(
            text=metadata_text,
            style=ExcelStyle(font_size=11),
            start_row=5
        )
        
        self.components["Summary"].extend([title_component, metadata_component])
        
    def create_worksheet(self, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create a new worksheet."""
        if name in self.worksheets:
            return self.worksheets[name]
            
        worksheet = self.workbook.create_sheet(name)
        self.worksheets[name] = worksheet
        self.components[name] = []
        return worksheet
        
    def add_component(self, component: ExcelComponent, worksheet_name: str = "Summary") -> None:
        """Add component to specific worksheet."""
        if worksheet_name not in self.worksheets:
            self.create_worksheet(worksheet_name)
            
        self.components[worksheet_name].append(component)
        
    def add_text(self, text: str, worksheet_name: str = "Summary", **kwargs) -> TextComponent:
        """Add text component to worksheet."""
        component = TextComponent(text=text, **kwargs)
        self.add_component(component, worksheet_name)
        return component
        
    def add_dataframe(self, df: pd.DataFrame, worksheet_name: str = "Data", **kwargs) -> DataFrameComponent:
        """Add DataFrame component to worksheet."""
        component = DataFrameComponent(df=df, **kwargs)
        self.add_component(component, worksheet_name)
        return component
        
    def add_plotly_chart(self, figure: go.Figure | None = None, worksheet_name: str = "Charts", **kwargs) -> PlotlyChartComponent:
        """Add Plotly chart component to worksheet."""
        component = PlotlyChartComponent(figure=figure, **kwargs)
        self.add_component(component, worksheet_name)
        return component
        
    def generate_report(self) -> str:
        """Generate the complete Excel report."""
        # Write all components to their respective worksheets
        for ws_name, components in self.components.items():
            worksheet = self.worksheets[ws_name]
            current_row = 1
            
            for component in components:
                current_row = component.write_to_worksheet(worksheet, current_row)
                
        # Save workbook
        self.workbook.save(self.filename)
        return self.filename
        
    def save(self) -> str:
        """Save the Excel report."""
        return self.generate_report()


# Helper functions for common quant formatting
def get_financial_number_formats() -> dict[str, str]:
    """Get common financial number formats tailored for quant reports."""
    return {
        'currency': '"$"#,##0.00',
        'currency_millions': '"$"#,##0,,"M"',
        'percentage': '0.00%',
        'basis_points': '0"bp"',
        'ratio': '0.00',
        'large_number': '#,##0',
        'return_pct': '0.00"%"',
        'pnl': '"$"#,##0.00_);[Red]("$"#,##0.00)',
        'volatility': '0.00%',
        'sharpe': '0.00'
    }


def create_pnl_conditional_formatting() -> dict[str, str]:
    """Get P&L specific conditional formatting rules."""
    return {
        'type': 'positive_negative'
    }


def create_sample_quant_excel_report() -> ExcelReport:
    """Create a comprehensive sample Excel report for quant teams."""
    
    # Create report
    report = ExcelReport("daily_quant_report.xlsx", "Daily Quantitative Analysis Report")
    
    # Add summary text
    report.add_text(
        """
Daily Performance Summary:
• Portfolio returned +1.87% today
• Outperformed benchmark by +0.92%
• All risk limits maintained within targets
• Strong performance in equity long/short strategy
    """, worksheet_name="Summary", title="Executive Summary")
    
    # Create sample P&L data
    np.random.seed(42)
    strategies = ['Equity Long/Short', 'Fixed Income Arb', 'Volatility Trading', 
                  'Market Neutral', 'Event Driven']
    
    pnl_data = pd.DataFrame({
        'Strategy': strategies,
        'Daily_PnL_MM': np.random.normal(1.5, 2.0, 5),
        'MTD_PnL_MM': np.random.normal(8.0, 12.0, 5),
        'YTD_PnL_MM': np.random.normal(45.0, 35.0, 5),
        'Daily_Return_Pct': np.random.normal(1.2, 1.8, 5),
        'Sharpe_Ratio': np.random.uniform(0.8, 2.5, 5),
        'Max_DD_Pct': -np.abs(np.random.normal(3.0, 2.0, 5)),
        'AUM_MM': np.random.uniform(50, 200, 5)
    })
    
    # Add P&L table with financial formatting
    financial_formats = get_financial_number_formats()
    pnl_formats = {
        'Daily_PnL_MM': financial_formats['pnl'],
        'MTD_PnL_MM': financial_formats['currency_millions'], 
        'YTD_PnL_MM': financial_formats['currency_millions'],
        'Daily_Return_Pct': financial_formats['percentage'],
        'Sharpe_Ratio': financial_formats['sharpe'],
        'Max_DD_Pct': financial_formats['percentage'],
        'AUM_MM': financial_formats['currency_millions']
    }
    
    pnl_conditional = {
        'Daily_PnL_MM': create_pnl_conditional_formatting(),
        'Daily_Return_Pct': {'type': 'color_scale'},
        'Sharpe_Ratio': {'type': 'data_bars'}
    }
    
    report.add_dataframe(
        pnl_data,
        worksheet_name="P&L Analysis", 
        title="Strategy Performance Breakdown",
        table_style=TableStyle.pnl_table(),
        number_formats=pnl_formats,
        conditional_formatting=pnl_conditional
    )
    
    # Create risk metrics table
    risk_data = pd.DataFrame({
        'Risk_Metric': ['VaR (99%, 1-day)', 'Expected Shortfall', 'Portfolio Beta', 
                       'Correlation to SPY', 'Max Drawdown', 'Volatility (Ann.)'],
        'Current_Value': [2.34, 3.12, 0.67, 0.74, -2.1, 14.2],
        'Limit_Target': [5.0, 4.0, 1.0, 0.8, -5.0, 18.0],
        'Utilization_Pct': [46.8, 78.0, 67.0, 92.5, 42.0, 78.9],
        'Status': ['OK', 'OK', 'OK', 'WATCH', 'OK', 'OK']
    })
    
    risk_formats = {
        'Current_Value': financial_formats['ratio'],
        'Limit_Target': financial_formats['ratio'],
        'Utilization_Pct': financial_formats['percentage']
    }
    
    risk_conditional = {
        'Utilization_Pct': {'type': 'color_scale'},
    }
    
    report.add_dataframe(
        risk_data,
        worksheet_name="Risk Metrics",
        title="Daily Risk Dashboard", 
        table_style=TableStyle.risk_table(),
        number_formats=risk_formats,
        conditional_formatting=risk_conditional
    )
    
    # Add Plotly charts if available
    if HAS_PLOTLY:
        # Cumulative returns chart
        dates = pd.date_range('2024-01-01', '2024-08-09', freq='D')
        portfolio_returns = np.random.normal(0.0008, 0.015, len(dates))
        benchmark_returns = np.random.normal(0.0005, 0.012, len(dates))
        
        portfolio_cum = (1 + pd.Series(portfolio_returns, index=dates)).cumprod()
        benchmark_cum = (1 + pd.Series(benchmark_returns, index=dates)).cumprod()
        
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=dates, y=(portfolio_cum-1)*100, name='Portfolio', 
                                 line=dict(color='blue', width=2)))
        fig.add_trace(go.Scatter(x=dates, y=(benchmark_cum-1)*100, name='Benchmark', 
                                 line=dict(color='gray', width=1, dash='dash')))
        
        fig.update_layout(
            title="Cumulative Returns Comparison",
            xaxis_title="Date",
            yaxis_title="Cumulative Return (%)",
            template="plotly_white",
            height=400
        )
        
        report.add_plotly_chart(fig, worksheet_name="Charts", title="Performance Chart", height=400)
        
        # Risk decomposition pie chart
        risk_contrib = pd.DataFrame({
            'Strategy': strategies[:4],  # Top 4 for pie chart
            'Risk_Contribution': [35.2, 28.1, 22.7, 14.0]
        })
        
        fig2 = px.pie(risk_contrib, values='Risk_Contribution', names='Strategy', 
                      title='Risk Contribution by Strategy')
        fig2.update_traces(textposition='inside', textinfo='percent+label')
        fig2.update_layout(height=400)
        
        report.add_plotly_chart(fig2, worksheet_name="Charts", title="Risk Decomposition", height=400)
    
    return report


if __name__ == "__main__":
    # Create and generate sample Excel report
    sample_report = create_sample_quant_excel_report()
    filename = sample_report.save()
    print(f"Excel report generated: {filename}")
